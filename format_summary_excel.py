#!/usr/bin/env python3
"""Format existing DOF batch summary Excel files."""

from __future__ import annotations

import argparse
import math
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def _truncated_numeric_value(value: Any) -> tuple[Any, bool]:
    if value is None:
        return value, False
    if isinstance(value, bool):
        return value, False
    if isinstance(value, int):
        return value, False
    if isinstance(value, float):
        if not math.isfinite(value):
            return None, False
        return math.trunc(value * 100000.0) / 100000.0, True
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return value, False
        if text.lower() in {"true", "false"}:
            return value, False
        try:
            number = float(text)
        except ValueError:
            return value, False
        if not math.isfinite(number):
            return None, False
        if re.fullmatch(r"[+-]?\d+", text):
            return int(text), False
        return math.trunc(number * 100000.0) / 100000.0, True
    return value, False


def format_summary_excel(input_path: Path, output_path: Path | None = None, overwrite: bool = False) -> Path:
    input_path = input_path.expanduser().resolve()
    if not input_path.exists():
        raise FileNotFoundError(f"Excel file not found: {input_path}")
    if output_path is None:
        output_path = input_path if overwrite else input_path.with_name(f"{input_path.stem}_formatted{input_path.suffix}")
    else:
        output_path = output_path.expanduser().resolve()
    if output_path == input_path and not overwrite:
        raise ValueError("Refusing to overwrite input without --overwrite")

    wb = load_workbook(input_path)
    changed = 0
    for ws in wb.worksheets:
        header_to_col = {
            str(cell.value).strip(): cell.column
            for cell in ws[1]
            if cell.value is not None
        }
        case_name_col = header_to_col.get("case_name", 1)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if cell.column == case_name_col:
                    continue
                new_value, is_float = _truncated_numeric_value(cell.value)
                if new_value != cell.value:
                    cell.value = new_value
                    changed += 1
                if is_float:
                    cell.number_format = "0.00000"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"wrote: {output_path}")
    print(f"changed_cells: {changed}")
    return output_path


def main() -> None:
    parser = argparse.ArgumentParser(description="Truncate numeric cells in an existing summary Excel to 5 decimals.")
    parser.add_argument(
        "input",
        nargs="?",
        default="new_dof_batch_results/batch_summary.xlsx",
        help="Input .xlsx path. Default: new_dof_batch_results/batch_summary.xlsx",
    )
    parser.add_argument("-o", "--output", help="Output .xlsx path. Default: <input>_formatted.xlsx")
    parser.add_argument("--overwrite", action="store_true", help="Overwrite input file in place.")
    args = parser.parse_args()

    format_summary_excel(
        Path(args.input),
        Path(args.output) if args.output else None,
        overwrite=bool(args.overwrite),
    )


if __name__ == "__main__":
    main()
